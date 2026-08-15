"""本地离线 Web 校对工作台（127.0.0.1，零外网依赖）。

特性：
- 全程 OfflineGuard 封锁（仅回环可用）；
- 分块上传落盘；识别为异步任务（进度/取消可查）；
- 页面图片按需渲染并缓存（大 PDF 友好）；
- 全文级富文本编辑（Quill2 本地化）+ 编辑后导出（DOCX/TXT/双层PDF）。
"""
from __future__ import annotations

import json
import logging
import mimetypes
import os
import re
import threading
import uuid

log = logging.getLogger("mdun.web")
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse

from mdun.config import load_settings
from mdun.pipeline import Pipeline, ProcessingCancelled
from mdun.export import export_txt, export_markdown, export_json
from mdun.export.pdf import export_searchable_pdf
from mdun.export.ops_docx import export_docx_from_ops
from mdun.postprocess import repair_punctuation
from mdun.security import OfflineGuard, AuditLog

STATIC_DIR = Path(__file__).resolve().parent / "static"
_store: dict[str, dict] = {}
_store_lock = threading.Lock()
_store_dir: Path | None = None          # 项目存档目录（支持用户指定路径）


def _storage_settings_path() -> Path:
    return Path(_settings.data_dir) / "storage.json"


def _resolve_store_dir() -> Path:
    global _store_dir
    if _store_dir is not None:
        return _store_dir
    _store_dir = Path(_settings.data_dir) / "projects"
    try:
        import json as _json

        sp = _storage_settings_path()
        if sp.exists():
            saved = _json.loads(sp.read_text(encoding="utf-8"))
            cand = Path(saved.get("path", "")).expanduser()
            if saved.get("path") and cand.is_absolute():
                cand.mkdir(parents=True, exist_ok=True)
                _store_dir = cand
    except Exception:  # noqa: BLE001
        pass
    _store_dir.mkdir(parents=True, exist_ok=True)
    return _store_dir


def _load_store() -> dict[str, dict]:
    out: dict[str, dict] = {}
    d = _resolve_store_dir()
    for f in d.glob("*.json"):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
            out[data.get("id") or f.stem] = data
        except Exception:  # noqa: BLE001
            continue
    return out


def _save_project(proj_id: str) -> None:
    with _store_lock:
        proj = _store.get(proj_id)
        if proj is None:
            return
        d = _resolve_store_dir()
        tmp = d / f"{proj_id}.json.tmp"
        tmp.write_text(json.dumps(proj, ensure_ascii=False), encoding="utf-8")
        tmp.replace(d / f"{proj_id}.json")


def _delete_project(proj_id: str | None) -> bool:
    """删除单个项目（内存/存档文件/页面缓存），供单个与批量删除共用。"""
    if not proj_id:
        return False
    with _store_lock:
        proj = _store.pop(proj_id, None)
    if proj is None:
        return False
    import shutil

    cache = Path(_settings.data_dir) / PAGE_CACHE_DIRNAME / proj_id
    shutil.rmtree(cache, ignore_errors=True)
    try:
        (_resolve_store_dir() / f"{proj_id}.json").unlink(missing_ok=True)
    except Exception:  # noqa: BLE001
        pass
    return True
_jobs: dict[str, dict] = {}
_jobs_lock = threading.Lock()
_worker = None
_settings = None
_t2s = None

MAX_UPLOAD = 2 * 1024 * 1024 * 1024  # 2GB
PAGE_CACHE_DIRNAME = "page_cache"
MAX_RENDER_SIDE = 1800  # 页面图片最长边

class Handler(BaseHTTPRequestHandler):
    server_version = "OctoOCR/0.3"

    def log_message(self, fmt, *args):  # 访问日志静默（审计走 AuditLog）
        pass

    # ---- helpers ----
    def _send(self, code: int, body: bytes, ctype: str = "application/json; charset=utf-8",
              extra_headers: dict | None = None) -> None:
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        for k, v in (extra_headers or {}).items():
            self.send_header(k, v)
        self.end_headers()
        try:
            self.wfile.write(body)
        except BrokenPipeError:
            pass

    def _json(self, obj, code: int = 200) -> None:
        self._send(code, json.dumps(obj, ensure_ascii=False).encode("utf-8"))

    def _read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", 0))
        try:
            return json.loads(self.rfile.read(length) or b"{}")
        except Exception:  # noqa: BLE001
            return {}

    # ---- routes ----
    def do_GET(self):
        url = urlparse(self.path)
        if url.path == "/":
            return self._static("index.html")
        if url.path.startswith("/static/"):
            return self._static(url.path[len("/static/"):])
        m = re.match(r"^/api/project/([^/]+)$", url.path)
        if m:
            with _store_lock:
                proj = _store.get(m.group(1))
            return self._json(proj) if proj else self._json({"error": "not found"}, 404)
        m = re.match(r"^/api/page_image/([^/]+)/(\d+)$", url.path)
        if m:
            return self._page_image(m.group(1), int(m.group(2)))
        m = re.match(r"^/api/job/([^/]+)$", url.path)
        if m:
            return self._job_status(m.group(1))
        m = re.match(r"^/api/download/([^/]+)$", url.path)
        if m:
            return self._handle_download(m.group(1))
        if url.path == "/api/list":
            with _store_lock:
                return self._json([
                    {"id": k, "source": v.get("filename") or Path(v["source"]).name, "pages": len(v["pages"])}
                    for k, v in _store.items()
                ])
        return self._json({"error": "not found"}, 404)

    def do_POST(self):
        url = urlparse(self.path)
        if url.path == "/api/ocr":
            return self._handle_upload()
        if url.path == "/api/repair":
            return self._handle_repair()
        if url.path == "/api/save_edit":
            return self._handle_save_edit()
        if url.path == "/api/detect_regions":
            return self._handle_detect_regions()
        if url.path == "/api/spellcheck":
            return self._handle_spellcheck()
        if url.path == "/api/region_mode":
            return self._handle_region_mode()
        if url.path == "/api/recognize_region":
            return self._handle_recognize_region()
        if url.path == "/api/reprocess":
            return self._handle_reprocess()
        if url.path == "/api/delete":
            return self._handle_delete()
        if url.path == "/api/delete_many":
            return self._handle_delete_many()
        if url.path == "/api/pick_folder":
            return self._handle_pick_folder()
        if url.path == "/api/t2s":
            return self._handle_t2s()
        if url.path == "/api/storage":
            return self._handle_storage()
        if url.path == "/api/export":
            return self._handle_export()
        if url.path.startswith("/api/cancel/"):
            return self._handle_cancel(url.path.rsplit("/", 1)[-1])
        return self._json({"error": "not found"}, 404)

    # ---- static ----
    def _static(self, name: str) -> None:
        f = (STATIC_DIR / name).resolve()
        if not f.is_file() or STATIC_DIR.resolve() not in f.parents:
            return self._json({"error": "not found"}, 404)
        ctype = mimetypes.guess_type(str(f))[0] or "application/octet-stream"
        self._send(200, f.read_bytes(), ctype)

    # ---- page image（懒渲染 + 磁盘缓存，大 PDF 友好）----
    def _page_image(self, proj_id: str, page_idx: int) -> None:
        with _store_lock:
            proj = _store.get(proj_id)
        if not proj:
            return self._json({"error": "not found"}, 404)
        cache_dir = Path(_settings.data_dir) / PAGE_CACHE_DIRNAME / proj_id
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_file = cache_dir / f"p{page_idx}.png"
        if not cache_file.exists():
            try:
                png = self._render_page_png(proj["source"], page_idx)
                cache_file.write_bytes(png)
            except Exception as e:  # noqa: BLE001
                return self._json({"error": f"页面渲染失败: {e}"}, 500)
        self._send(200, cache_file.read_bytes(), "image/png")

    @staticmethod
    def _render_page_png(source: str, page_idx: int) -> bytes:
        src = Path(source)
        if src.suffix.lower() == ".pdf":
            import pymupdf

            doc = pymupdf.open(str(src))
            try:
                if page_idx >= len(doc):
                    raise IndexError("page out of range")
                page = doc[page_idx]
                r = page.rect
                zoom = min(2.5, MAX_RENDER_SIDE / max(r.width, r.height))
                zoom = max(0.5, zoom)
                pix = page.get_pixmap(matrix=pymupdf.Matrix(zoom, zoom), alpha=False)
                return pix.tobytes("png")
            finally:
                doc.close()
        else:
            from PIL import Image
            import io

            with Image.open(src) as im:
                im = im.convert("RGB")
                im.thumbnail((MAX_RENDER_SIDE, MAX_RENDER_SIDE))
                buf = io.BytesIO()
                im.save(buf, "PNG")
                return buf.getvalue()

    # ---- 上传（分块落盘）+ 异步识别任务 ----
    def _handle_upload(self) -> None:
        ctype = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in ctype:
            return self._json({"error": "需要 multipart/form-data"}, 400)
        length = int(self.headers.get("Content-Length", 0))
        if length > MAX_UPLOAD:
            return self._json({"error": "文件超过 2GB 上限"}, 413)
        boundary = ctype.split("boundary=")[1].strip().strip('"').encode()
        tmp = _settings.data_dir / "tmp"
        tmp.mkdir(parents=True, exist_ok=True)
        upload_id = uuid.uuid4().hex[:10]

        # 读取 multipart 头（含 boundary 首行），文件体流式落盘
        # 注意：必须按 Content-Length 截断读取，读至 EOF 会与 keep-alive 客户端死锁
        chunk = 1 << 20
        remaining = length
        buf = bytearray()
        while remaining > 0 and boundary + b"\r\n" not in buf and len(buf) < 1 << 20:
            piece = self.rfile.read(min(chunk, remaining))
            if not piece:
                break
            buf += piece
            remaining -= len(piece)
        head, sep, rest = bytes(buf).partition(boundary + b"\r\n")
        if not sep:
            return self._json({"error": "multipart 解析失败"}, 400)
        disp, _, body_head = rest.partition(b"\r\n\r\n")
        m = re.search(rb'filename="([^"]+)"', disp)
        filename = m.group(1).decode("utf-8", "ignore") if m else "upload.bin"
        ext = Path(filename).suffix.lower()
        if ext not in (".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"):
            return self._json({"error": f"不支持的格式: {ext}"}, 400)
        save = tmp / f"upload_{upload_id}{ext}"
        tail_keep = len(boundary) + 16
        with open(save, "wb") as f:
            f.write(body_head)
            tail_buf = b""
            while remaining > 0:
                piece = self.rfile.read(min(chunk, remaining))
                if not piece:
                    break
                remaining -= len(piece)
                f.write(piece)
                tail_buf = (tail_buf + piece)[-tail_keep:]
        # 截掉末尾 "--boundary--\r\n"
        with open(save, "r+b") as f:
            f.seek(0, os.SEEK_END)
            size = f.tell()
            f.seek(max(0, size - tail_keep))
            tail = f.read()
            cut = tail.rfind(b"--" + boundary)
            if cut >= 0:
                new_size = size - tail_keep + cut - 2  # -2: boundary 前的 \r\n
                f.truncate(max(0, new_size))

        job_id = uuid.uuid4().hex[:12]
        try:
            from mdun.pipeline import _count_pages

            total_pages = _count_pages(str(save))
        except Exception:  # noqa: BLE001
            total_pages = 0
        with _jobs_lock:
            _jobs[job_id] = {
                "id": job_id, "status": "queued",
                "progress": {"done": 0, "total": total_pages},
                "file": str(save.resolve()), "filename": filename, "error": None,
                "coarse": True,  # 上传先粗识别，用户可再选全文识别
            }
        _ensure_worker()
        return self._json({"job_id": job_id, "status": "queued"})

    def _job_status(self, job_id: str) -> None:
        with _jobs_lock:
            job = _jobs.get(job_id)
            if not job:
                return self._json({"error": "not found"}, 404)
            resp = {k: v for k, v in job.items() if k != "file"}
            return self._json(resp)

    def _handle_cancel(self, job_id: str) -> None:
        with _jobs_lock:
            job = _jobs.get(job_id)
            if not job:
                return self._json({"error": "not found"}, 404)
            if job["status"] in ("running", "queued"):
                job["cancel_requested"] = True
                job["status"] = "canceling"
            return self._json({"status": job["status"]})

    # ---- 修复 ----
    def _handle_repair(self) -> None:
        req = self._read_json()
        text = req.get("text", "")
        new_text, edits = repair_punctuation(text)
        return self._json({"text": new_text, "edits": [e.__dict__ for e in edits]})

    # ---- 繁体转简体（opencc，纯本地）----
    def _handle_t2s(self) -> None:
        global _t2s
        req = self._read_json()
        texts = req.get("texts") or []
        if _t2s is None:
            from opencc import OpenCC

            _t2s = OpenCC("t2s")
        try:
            out = [_t2s.convert(t) for t in texts]
        except Exception as e:  # noqa: BLE001
            return self._json({"error": f"转换失败: {e}"}, 500)
        return self._json({"texts": out})

    # ---- 编辑保存（Quill ops）----
    def _handle_save_edit(self) -> None:
        req = self._read_json()
        proj_id = req.get("id")
        ops = req.get("ops")
        with _store_lock:
            proj = _store.get(proj_id)
        if not proj:
            return self._json({"error": "not found"}, 404)
        pages, plain_per_page = _ops_to_pages(ops, proj.get("pages") or [])
        proj["edited"] = True
        proj["edited_ops"] = ops
        proj["pages"] = pages
        proj["edited_page_texts"] = plain_per_page
        _save_project(proj_id)
        return self._json({"saved": True, "pages": len(pages)})

    # ---- 局部重新识别（框选区域 → 表格/公式）----
    def _handle_recognize_region(self) -> None:
        global _region_pipe
        req = self._read_json()
        proj_id = req.get("id")
        with _store_lock:
            proj = _store.get(proj_id)
        if not proj:
            return self._json({"error": "not found"}, 404)
        page_idx = int(req.get("page", 0))
        box = req.get("box") or [0.1, 0.1, 0.9, 0.9]
        pages = proj.get("pages") or []
        if page_idx >= len(pages):
            return self._json({"error": "page out of range"}, 404)
        img = _page_numpy(proj["source"], page_idx, dpi=300)
        if img is None:
            return self._json({"error": "页面渲染失败"}, 500)
        h, w = img.shape[:2]
        x0, y0 = int(box[0] * w), int(box[1] * h)
        x1, y1 = int(box[2] * w), int(box[3] * h)
        crop = img[max(0, y0):y1, max(0, x0):x1]
        if crop.size == 0:
            return self._json({"error": "区域无效"}, 400)
        with _region_lock:
            if _region_pipe is None:
                from mdun.pipeline import Pipeline

                _region_pipe = Pipeline(_settings)
            pipe = _region_pipe
        from mdun.pipeline import _sane_latex

        tables = 0
        formulas = 0
        latex = ""
        if pipe.table_engine.available:
            try:
                # 版面先行校验（页级）：用户框与版面模型检出的表格区域双向重叠 ≥30%
                # 才走表格结构识别，否则 SLANet 会把整段文本强制切成伪表格
                table_ok = True
                table_crop = crop
                table_box = [x0, y0, x1, y1]
                if getattr(pipe, "layout_engine", None) and pipe.layout_engine.available:
                    from mdun.ocr.region import TABLE_CLASSES as _TC

                    lregs = pipe.layout_engine.predict(img)
                    box_area = max(1.0, float((x1 - x0) * (y1 - y0)))
                    table_ok = False
                    for lr in lregs:
                        if lr.type not in _TC or lr.score < 0.4:
                            continue
                        bx0, by0, bx1, by1 = lr.box
                        ix = max(0.0, min(x1, bx1) - max(x0, bx0))
                        iy = max(0.0, min(y1, by1) - max(y0, by0))
                        inter = ix * iy
                        tarea = max(1.0, float((bx1 - bx0) * (by1 - by0)))
                        if inter / box_area >= 0.3 or inter / tarea >= 0.3:
                            table_ok = True
                            rx0, ry0 = max(x0, bx0), max(y0, by0)
                            rx1, ry1 = min(x1, bx1), min(y1, by1)
                            if rx1 - rx0 > 20 and ry1 - ry0 > 20:
                                table_crop = img[ry0:ry1, rx0:rx1]
                                table_box = [rx0, ry0, rx1, ry1]
                            break
                if table_ok:
                    crops, idx = pipe.table_engine.cell_crops(table_crop)
                    texts = pipe.engine.recognize_crops(crops)
                    n_rows = max((i for i, _ in idx), default=-1) + 1
                    n_cols = max((j for _, j in idx), default=-1) + 1
                    if n_cols >= 2:
                        rows = [[""] * n_cols for _ in range(max(n_rows, 1))]
                        for (ri, ci), t in zip(idx, texts):
                            if 0 <= ri < len(rows) and 0 <= ci < n_cols:
                                rows[ri][ci] = (rows[ri][ci] + " " + t).strip()
                        if n_rows == 1 and n_cols >= 3 and max(len(c) for c in rows[0]) <= 3:
                            pass  # 单行且各列极短 → 标题行被切碎的伪表格
                        elif sum(len(c) for r in rows for c in r) >= 20:
                            pages[page_idx].setdefault("tables", []).append({
                                "box": table_box, "rows": rows,
                                "html": "", "source": "region", "score": 1.0,
                            })
                            tables = 1
            except Exception as e:  # noqa: BLE001
                log.warning("区域表格识别失败: %s", e)
        # 文字与公式：文字先行。公式引擎会把文字区域幻觉成公式（同区域两次结果不一致），
        # 因此以 OCR 结果为准：CJK 充分 → 文字；文字不足且公式候选通过校验 → 公式。
        region_text = ""
        latex_cand = None
        if tables == 0:
            try:
                _, region_text = pipe.engine.recognize(crop)
                region_text = region_text.strip()
            except Exception as e:  # noqa: BLE001
                log.warning("区域文字识别失败: %s", e)
        if tables == 0 and pipe.formula_engine.available:
            try:
                f = pipe.formula_engine.recognize(crop, (x0, y0, x1, y1))
                if _sane_latex(f.latex):
                    latex_cand = f
            except Exception as e:  # noqa: BLE001
                log.warning("区域公式识别失败: %s", e)
        cjk = sum(1 for c in region_text if "\u4e00" <= c <= "\u9fff")
        if tables == 0:
            if cjk >= 4:
                # 文字区域：去重后并入该页正文
                p = pages[page_idx]
                if region_text not in (p.get("text") or ""):
                    pw = float(p.get("width") or w)
                    ph = float(p.get("height") or h)
                    sx, sy = (pw / w) if w else 1.0, (ph / h) if h else 1.0
                    p.setdefault("paras", []).append({
                        "kind": "text", "text": region_text,
                        "box": [x0 * sx, y0 * sy, x1 * sx, y1 * sy],
                    })
                    p["text"] = (p.get("text") or "").rstrip() + "\n\n" + region_text
            elif latex_cand is not None:
                pages[page_idx].setdefault("formulas", []).append({
                    "box": [x0, y0, x1, y1], "latex": latex_cand.latex,
                    "score": round(latex_cand.score, 3), "source": "region",
                })
                formulas = 1
                latex = latex_cand.latex
                region_text = ""
            elif len(region_text) >= 2:
                p = pages[page_idx]
                if region_text not in (p.get("text") or ""):
                    p.setdefault("paras", []).append({
                        "kind": "text", "text": region_text, "box": [x0, y0, x1, y1],
                    })
                    p["text"] = (p.get("text") or "").rstrip() + "\n\n" + region_text
            else:
                region_text = ""
        _save_project(proj_id)
        return self._json({"tables": tables, "formulas": formulas, "latex": latex, "text": region_text})

    # ---- 区域模式（全文 / 跳过选区 / 仅识别选区）----
    def _handle_region_mode(self) -> None:
        req = self._read_json()
        proj_id = req.get("id")
        with _store_lock:
            proj = _store.get(proj_id)
        if not proj:
            return self._json({"error": "not found"}, 404)
        pages = proj.get("pages") or []
        mode = req.get("mode", "full")
        apply_all = bool(req.get("apply_all"))
        boxes = req.get("boxes") or []
        norm = []
        for b in boxes:
            if isinstance(b, dict):
                x0, y0, x1, y1 = b.get("x0", 0), b.get("y0", 0), b.get("x1", 0), b.get("y1", 0)
            else:
                x0, y0, x1, y1 = b[0], b[1], b[2], b[3]
            x0, y0 = max(0.0, min(1.0, float(x0))), max(0.0, min(1.0, float(y0)))
            x1, y1 = max(0.0, min(1.0, float(x1))), max(0.0, min(1.0, float(y1)))
            if x1 > x0 and y1 > y0:
                norm.append({"x0": x0, "y0": y0, "x1": x1, "y1": y1})
        if mode != "full" and not norm:
            return self._json({"error": "未提供标注区域"}, 400)
        target_idx = [int(req.get("page", 0))] if not apply_all else list(range(len(pages)))
        removed_total = 0
        for i in target_idx:
            if i >= len(pages):
                continue
            p = pages[i]
            if mode == "full":
                p["region_mode"] = "full"
                p["ignore_regions"] = []
                removed_total += _restore_full(p)
                continue
            p["region_mode"] = mode
            p["ignore_regions"] = norm
            removed_total += _apply_region_mode(proj, p, mode, norm)
        _save_project(proj_id)
        return self._json({"saved": True, "mode": mode, "pages": len(target_idx), "removed": removed_total})

    # ---- 全文重识别（粗识别后升级精度/表格/公式）----
    def _handle_reprocess(self) -> None:
        req = self._read_json()
        proj_id = req.get("id")
        with _store_lock:
            proj = _store.get(proj_id)
        if not proj:
            return self._json({"error": "not found"}, 404)
        job_id = uuid.uuid4().hex[:12]
        with _jobs_lock:
            _jobs[job_id] = {
                "id": job_id, "status": "queued",
                "progress": {"done": 0, "total": 0},
                "file": proj["source"], "filename": proj.get("filename"),
                "error": None, "coarse": False, "reprocess": proj_id,
            }
        _ensure_worker()
        return self._json({"job_id": job_id, "status": "queued"})

    def _handle_delete(self) -> None:
        req = self._read_json()
        if not _delete_project(req.get("id")):
            return self._json({"error": "not found"}, 404)
        return self._json({"deleted": True})

    # ---- 批量删除（任务区多选）----
    def _handle_delete_many(self) -> None:
        req = self._read_json()
        ids = [str(i) for i in (req.get("ids") or [])][:500]
        deleted = sum(1 for pid in ids if _delete_project(pid))
        return self._json({"deleted": deleted})

    # ---- 拼写检查 ----
    def _handle_spellcheck(self) -> None:
        global _spell_checker
        req = self._read_json()
        text = req.get("text", "")
        if _spell_checker is None:
            from mdun.postprocess.spell import SpellChecker

            _spell_checker = SpellChecker()
        try:
            items = _spell_checker.check(text)
        except Exception as e:  # noqa: BLE001
            return self._json({"error": f"拼写检查失败: {e}"}, 500)
        return self._json({
            "items": [
                {"start": it.start, "end": it.end, "word": it.word,
                 "suggestions": it.suggestions, "lang": it.lang}
                for it in items
            ],
        })

    def _handle_download(self, token: str) -> None:
        with _dl_lock:
            path = _downloads.get(token)
        if not path or not Path(path).exists():
            return self._json({"error": "下载链接不存在或已过期"}, 404)
        from urllib.parse import quote

        name = Path(path).name
        headers = {
            "Content-Disposition": "attachment; filename*=UTF-8''" + quote(name),
        }
        ctype = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        self._send(200, Path(path).read_bytes(), ctype, extra_headers=headers)

    # ---- 特殊格式识别（手动触发：低阈值版面 + 表格/公式识别）----
    def _handle_detect_regions(self) -> None:
        global _region_pipe
        req = self._read_json()
        proj_id = req.get("id")
        page_idx = int(req.get("page", 0))
        with _store_lock:
            proj = _store.get(proj_id)
        if not proj:
            return self._json({"error": "not found"}, 404)
        pages = proj.get("pages") or []
        if page_idx >= len(pages):
            return self._json({"error": "page out of range"}, 404)

        img = _page_numpy(proj["source"], page_idx)
        if img is None:
            return self._json({"error": "页面渲染失败"}, 500)
        with _region_lock:
            if _region_pipe is None:
                from mdun.pipeline import Pipeline

                _region_pipe = Pipeline(_settings)
            pipe = _region_pipe
        if not pipe.layout_engine.available:
            return self._json({"error": "版面模型缺失", "candidates": []}, 400)

        from mdun.ocr.region import TABLE_CLASSES, FORMULA_CLASSES
        from mdun.pipeline import _sane_latex

        regions = pipe.layout_engine.predict(img, low=True)
        candidates = [
            {"type": r.type, "score": round(r.score, 3), "box": [int(v) for v in r.box]}
            for r in regions
            if r.type in TABLE_CLASSES or r.type in FORMULA_CLASSES or r.type == "figure"
        ]
        tables: list[dict] = []
        formulas: list[dict] = []
        h, w = img.shape[:2]
        page_area = h * w
        if pipe.table_engine.available:
            for r in regions:
                if r.type not in TABLE_CLASSES:
                    continue
                area = (r.box[2] - r.box[0]) * (r.box[3] - r.box[1])
                if area > page_area * 0.85:
                    continue
                crop = img[max(0, r.box[1]):r.box[3], max(0, r.box[0]):r.box[2]]
                if crop.size == 0:
                    continue
                try:
                    crops, idx = pipe.table_engine.cell_crops(crop)
                    texts = pipe.engine.recognize_crops(crops)
                    n_rows = max((i for i, _ in idx), default=-1) + 1
                    n_cols = max((j for _, j in idx), default=-1) + 1
                    if n_cols < 2:
                        continue
                    rows = [[""] * n_cols for _ in range(max(n_rows, 1))]
                    for (ri, ci), t in zip(idx, texts):
                        if 0 <= ri < len(rows) and 0 <= ci < n_cols:
                            rows[ri][ci] = (rows[ri][ci] + " " + t).strip()
                    if n_rows == 1 and n_cols >= 3 and max(len(c) for c in rows[0]) <= 3:
                        continue
                    if sum(len(c) for r in rows for c in r) < 20:
                        continue
                    tables.append({
                        "box": [int(v) for v in r.box], "rows": rows,
                        "html": "", "source": "slanet_plus", "score": round(r.score, 3),
                    })
                except Exception as e:  # noqa: BLE001
                    log.warning("手动表格识别失败 %s: %s", r.box, e)
        if pipe.formula_engine.available:
            for r in regions:
                is_figure = r.type == "figure"
                if r.type not in FORMULA_CLASSES and not is_figure:
                    continue
                crop = img[max(0, r.box[1]):r.box[3], max(0, r.box[0]):r.box[2]]
                if crop.size == 0:
                    continue
                try:
                    f = pipe.formula_engine.recognize(crop, r.box)
                    if is_figure and not _sane_latex(f.latex):
                        continue
                    formulas.append({
                        "box": [int(v) for v in r.box], "latex": f.latex,
                        "score": round(f.score, 3), "source": f.source,
                    })
                except Exception as e:  # noqa: BLE001
                    log.warning("手动公式识别失败 %s: %s", r.box, e)

        # 合并回项目（页面表格/公式字段），并从正文段落中剔除重叠项
        pages[page_idx]["tables"] = tables
        pages[page_idx]["formulas"] = formulas
        skip = [tuple(t["box"]) for t in tables] + [tuple(f["box"]) for f in formulas]
        if skip:
            kept = []
            for para in pages[page_idx].get("paras", []):
                bx = para.get("box", [0, 0, 0, 0])
                cx, cy = (bx[0] + bx[2]) / 2, (bx[1] + bx[3]) / 2
                if any(x0 <= cx <= x1 and y0 <= cy <= y1 for x0, y0, x1, y1 in skip):
                    continue
                kept.append(para)
            pages[page_idx]["paras"] = kept
        return self._json({
            "page": page_idx,
            "candidates": candidates,
            "tables": len(tables),
            "formulas": len(formulas),
        })

    # ---- 原生文件夹选择对话框（存档位置用）----
    def _handle_pick_folder(self) -> None:
        import subprocess
        import sys

        picked = None
        try:
            if sys.platform == "darwin":
                script = 'POSIX path of (choose folder with prompt "选择存档文件夹")'
                out = subprocess.run(
                    ["osascript", "-e", script], capture_output=True, text=True, timeout=600
                )
                picked = out.stdout.strip().rstrip(":/")
            elif sys.platform.startswith("win"):
                ps = (
                    "Add-Type -AssemblyName System.Windows.Forms;"
                    "$d = New-Object System.Windows.Forms.FolderBrowserDialog;"
                    '$d.Description = "选择存档文件夹";'
                    'if ($d.ShowDialog() -eq [System.Windows.Forms.DialogResult]::OK) { Write-Output $d.SelectedPath }'
                )
                out = subprocess.run(
                    ["powershell", "-NoProfile", "-Command", ps], capture_output=True, text=True, timeout=600
                )
                picked = out.stdout.strip().rstrip("\\/")
            else:
                return self._json({"error": "此系统请直接输入路径"}, 400)
        except Exception as e:  # noqa: BLE001
            return self._json({"error": f"无法打开选择对话框: {e}"}, 500)
        if not picked:
            return self._json({"path": None, "canceled": True})
        return self._json({"path": picked})

    # ---- 存档位置（项目持久化目录，用户可指定）----
    def _handle_storage(self) -> None:
        req = self._read_json()
        if req.get("path"):
            cand = Path(req["path"]).expanduser()
            if not cand.is_absolute():
                return self._json({"error": "需要输入完整文件夹路径"}, 400)
            try:
                cand.mkdir(parents=True, exist_ok=True)
                # 迁移既有项目文件
                old = _resolve_store_dir()
                moved = 0
                if old.resolve() != cand.resolve():
                    for f in old.glob("*.json"):
                        if not (cand / f.name).exists():
                            f.replace(cand / f.name)
                            moved += 1
                global _store_dir
                _store_dir = cand
                _storage_settings_path().write_text(
                    json.dumps({"path": str(cand)}, ensure_ascii=False), encoding="utf-8"
                )
                return self._json({"path": str(cand), "moved": moved})
            except Exception as e:  # noqa: BLE001
                return self._json({"error": f"无法使用该文件夹: {e}"}, 400)
        d = _resolve_store_dir()
        with _store_lock:
            count = len(_store)
        return self._json({"path": str(d), "count": count})

    # ---- 导出 ----
    def _handle_export(self) -> None:
        req = self._read_json()
        proj_id = req.get("id")
        fmt = req.get("format", "txt")
        toc = req.get("toc") or None
        with _store_lock:
            proj = _store.get(proj_id)
        if not proj:
            return self._json({"error": "not found"}, 404)
        out_dir = _settings.data_dir / "exports"
        out_dir.mkdir(parents=True, exist_ok=True)
        stem = Path(proj.get("filename") or proj["source"]).stem

        if proj.get("edited_ops"):
            page_texts = proj.get("edited_page_texts") or [p["text"] for p in proj["pages"]]
            if fmt == "docx":
                f = export_docx_from_ops(proj["edited_ops"], out_dir / f"{stem}.edited.docx", toc=toc)
            elif fmt == "xlsx":
                from openpyxl import Workbook

                wb = Workbook()
                wb.remove(wb.active)
                n = 0
                for p in proj["pages"]:
                    for t in p.get("tables", []):
                        rows = t.get("rows", [])
                        if not rows:
                            continue
                        ws = wb.create_sheet(title=f"表{n + 1}-p{p['index'] + 1}")
                        for r in rows:
                            ws.append(r)
                        n += 1
                if n == 0:
                    wb.create_sheet(title="Sheet1")
                f = out_dir / f"{stem}.edited.xlsx"
                wb.save(str(f))
            elif fmt == "txt":
                f = out_dir / f"{stem}.edited.txt"
                f.write_text("\n\n".join(t for t in page_texts if t), encoding="utf-8")
            elif fmt == "pdf":
                f = export_searchable_pdf(proj["source"], out_dir / f"{stem}.edited.searchable.pdf", page_texts)
            elif fmt == "md":
                f = out_dir / f"{stem}.edited.md"
                f.write_text("\n\n".join(t for t in page_texts if t), encoding="utf-8")
            else:
                return self._json({"error": f"未知格式 {fmt}"}, 400)
            token = _register_download(f)
            return self._json({"file": str(f), "download": f"/api/download/{token}", "filename": Path(f).name})

        from mdun.pipeline import PageData, ParaOut, Project

        pages = []
        for p in proj["pages"]:
            paras = [ParaOut(kind=para["kind"], text=para["text"]) for para in p["paras"]]
            pages.append(PageData(index=p["index"], kind=p["kind"], text=p["text"], paras=paras))
        proj_obj = Project(source=proj["source"], engine_note=proj["engine"], pages=pages)
        if fmt == "txt":
            f = export_txt(proj_obj, out_dir / f"{stem}.txt")
        elif fmt == "md":
            f = export_markdown(proj_obj, out_dir / f"{stem}.md")
        elif fmt == "json":
            f = export_json(proj_obj, out_dir / f"{stem}.mdun.json")
        elif fmt == "pdf":
            f = export_searchable_pdf(proj["source"], out_dir / f"{stem}.searchable.pdf", [p["text"] for p in proj["pages"]])
        elif fmt == "docx":
            from mdun.export import export_docx

            f = export_docx(proj_obj, out_dir / f"{stem}.docx")
        else:
            return self._json({"error": f"未知格式 {fmt}"}, 400)
        token = _register_download(f)
        return self._json({"file": str(f), "download": f"/api/download/{token}", "filename": Path(f).name})

def _ops_to_pages(ops: list, orig_pages: list | None = None) -> tuple[list[dict], list[str]]:
    """Quill ops → 每页段落列表 + 每页纯文本（mdunPage embed 划分页面）。

    orig_pages 传入时保留其中不随编辑变化的结构字段（页面几何、行级识别结果、
    低置信行、忽略区域等），供图文对照与区域模式继续使用。
    """
    KEEP_PAGE_FIELDS = ("kind", "width", "height", "conf_avg", "seconds", "punc_edits",
                        "removed", "low_conf", "lines", "ignore_regions", "region_mode")
    pages: dict[int, list[dict]] = {}
    tables_by_page: dict[int, list[dict]] = {}
    seen_pages: set[int] = set()
    current_page = 0
    seen_pages.add(0)
    buf = ""
    cur_attrs: dict = {}

    def flush():
        nonlocal buf
        if buf.strip():
            kind = "heading" if cur_attrs.get("header") else "text"
            pages.setdefault(current_page, []).append({"kind": kind, "text": buf.strip(), "box": [0, 0, 0, 0]})
        buf = ""
        cur_attrs.clear()

    for op in ops:
        if not isinstance(op, dict):
            continue
        ins = op.get("insert")
        if isinstance(ins, str):
            buf += ins
            a = op.get("attributes") or {}
            if a:
                cur_attrs.update(a)
            continue
        if isinstance(ins, dict):
            if ins.get("mdunPage") is not None:
                flush()  # 先归档上一页，再切换页码
                current_page = int(ins["mdunPage"])
                seen_pages.add(current_page)
            elif ins.get("mdunTable") is not None:
                flush()
                tables_by_page.setdefault(current_page, []).append({
                    "box": [0, 0, 0, 0], "rows": ins["mdunTable"],
                    "html": "", "source": "edited", "score": 1.0,
                })
            else:
                flush()  # 其他 embed（图片等）结束当前段
    flush()
    # 页数由页面分隔 embed 决定：末页文本被清空时页结构不丢失
    max_page = max(seen_pages, default=0)
    orig_by_index = {int(p.get("index", i)): p for i, p in enumerate(orig_pages or [])}
    plain_per_page = ["\n\n".join(p["text"] for p in pages.get(i, [])) for i in range(max_page + 1)]
    out_pages: list[dict] = []
    for i in range(max_page + 1):
        page = {"index": i, "text": plain_per_page[i],
                "paras": pages.get(i, []), "tables": tables_by_page.get(i, [])}
        orig = orig_by_index.get(i)
        if orig:
            for k in KEEP_PAGE_FIELDS:
                if k in orig:
                    page[k] = orig[k]
        out_pages.append(page)
    return out_pages, plain_per_page

_region_pipe = None
_region_lock = threading.Lock()
_spell_checker = None
_downloads: dict[str, str] = {}
_dl_lock = threading.Lock()

def _register_download(path) -> str:
    import uuid

    token = uuid.uuid4().hex[:16]
    with _dl_lock:
        _downloads[token] = str(path)
    return token

def _page_numpy(source: str, page_idx: int, dpi: int = 200):
    """渲染单页图像为 RGB ndarray（手动特殊格式识别用）。"""
    import pymupdf
    import numpy as np

    if source.lower().endswith(".pdf"):
        doc = pymupdf.open(source)
        try:
            if page_idx >= len(doc):
                return None
            zoom = dpi / 72.0
            pix = doc[page_idx].get_pixmap(matrix=pymupdf.Matrix(zoom, zoom), alpha=False)
            img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
            return img[:, :, :3].copy()
        finally:
            doc.close()
    else:
        from PIL import Image

        with Image.open(source) as im:
            return np.asarray(im.convert("RGB"))

def _restore_full(page: dict) -> int:
    """恢复粗识别的全文（全文识别模式）。"""
    if "_full_text" in page:
        page["text"] = page.pop("_full_text")
        page["paras"] = page.pop("_full_paras", [])
        return 1
    return 0

def _apply_region_mode(proj: dict, page: dict, mode: str, regions: list) -> int:
    """exclude: 过滤框内文本；include: 仅保留框内文本。返回变化行数。"""
    w = float(page.get("width") or 0)
    h = float(page.get("height") or 0)
    if w <= 0 or h <= 0:
        return 0
    boxes = [(r["x0"] * w, r["y0"] * h, r["x1"] * w, r["y1"] * h) for r in regions]

    def inside(box, cx, cy):
        return box[0] <= cx <= box[2] and box[1] <= cy <= box[3]

    def any_inside(cx, cy):
        return any(inside(b, cx, cy) for b in boxes)

    # 备份原文（首次过滤前）
    if "_full_text" not in page:
        page["_full_text"] = page.get("text", "")
        page["_full_paras"] = page.get("paras", [])

    if page.get("kind") == "text":
        from mdun.ocr.document import extract_text_lines

        lines = extract_text_lines(proj["source"], int(page["index"]))
        kept = []
        for text, x0, y0, x1, y1 in lines:
            hit = any_inside((x0 + x1) / 2, (y0 + y1) / 2)
            if (mode == "exclude" and hit) or (mode == "include" and not hit):
                continue
            kept.append(text)
        page["paras"] = [{"kind": "text", "text": "\n".join(kept), "box": [0, 0, 0, 0]}] if kept else []
        page["text"] = "\n".join(kept)
        return len(lines) - len(kept)
    else:
        kept = []
        removed = 0
        for para in page.get("paras", []):
            bx = para.get("box", [0, 0, 0, 0])
            hit = any_inside((bx[0] + bx[2]) / 2, (bx[1] + bx[3]) / 2)
            if (mode == "exclude" and hit) or (mode == "include" and not hit):
                removed += 1
                continue
            kept.append(para)
        page["paras"] = kept
        page["text"] = "\n\n".join(p["text"] for p in kept)
        return removed
# ---- 后台识别任务（上传粗识别 / 全文重识别共用）----
_worker_lock = threading.Lock()


def _ensure_worker() -> None:
    global _worker
    with _worker_lock:
        if _worker is not None and _worker.is_alive():
            return
        _worker = threading.Thread(target=_worker_loop, name="octo-ocr-worker", daemon=True)
        _worker.start()


def _worker_loop() -> None:
    import time

    while True:
        job = None
        with _jobs_lock:
            for j in _jobs.values():
                if j.get("status") == "queued":
                    j["status"] = "running"
                    job = j
                    break
        if job is None:
            time.sleep(0.2)
            continue
        _run_job(job)


def _run_job(job: dict) -> None:
    """执行识别任务：粗识别（上传）与全文重识别（reprocess 替换原项目）共用。"""
    job_id = job["id"]
    src = job["file"]
    coarse = bool(job.get("coarse"))
    reprocess_id = job.get("reprocess")
    audit = AuditLog(Path(_settings.data_dir) / "audit.log")

    def on_progress(done: int, total: int, *_) -> None:
        with _jobs_lock:
            job["progress"] = {"done": done, "total": total}

    def cancel() -> bool:
        with _jobs_lock:
            return bool(job.get("cancel_requested"))

    try:
        pipe = Pipeline(_settings, audit=audit)
        project = pipe.process(
            src, dpi=200, use_punc_model=False, repair_punc=True, repair_para=True,  # 不自动补标点：修复本义是西文标点
            coarse=coarse, on_progress=on_progress, cancel=cancel,
        )
        from mdun.export.json import project_to_dict

        proj_id = reprocess_id or uuid.uuid4().hex[:12]
        data = project_to_dict(project)
        data["id"] = proj_id
        data["filename"] = job.get("filename") or Path(src).name
        with _store_lock:
            _store[proj_id] = data
        _save_project(proj_id)
        with _jobs_lock:
            job["status"] = "done"
            job["project_id"] = proj_id
            total = job["progress"].get("total", 0)
            job["progress"] = {"done": total, "total": total}
    except ProcessingCancelled:
        with _jobs_lock:
            job["status"] = "canceled"
    except Exception as e:  # noqa: BLE001
        log.exception("识别任务失败 %s", job_id)
        with _jobs_lock:
            job["status"] = "error"
            job["error"] = str(e)


def run(host: str = "127.0.0.1", port: int = 8788, data_dir: str | Path | None = None) -> None:
    """启动本地工作台：全程 OfflineGuard 封锁外网，仅回环可用。"""
    global _settings, _store
    _settings = load_settings(str(data_dir) if data_dir else None)
    _store = _load_store()
    OfflineGuard().enable()
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"OctoOCR 工作台已启动（完全离线）: http://{host}:{port}")
    print("按 Ctrl+C 退出。")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass

